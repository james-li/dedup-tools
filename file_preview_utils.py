import io
import os
import tempfile
import time
import traceback
import uuid

from docx2pdf import convert
from PIL import Image, ImageTk
import fitz
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from openpyxl import load_workbook


def convert_pdf_to_image(pdf_path):
    # return convert_from_path(pdf_path, poppler_path=r"D:\tools\poppler-0.68.0\bin")[0]
    # tmp_png = os.path.join(tempfile.gettempdir(), str(uuid.uuid4()) + "_" + str(int(time.time())) + ".png")
    try:
        doc = fitz.open(pdf_path)
        for page in doc:
            pix = page.get_pixmap()
            img_bytes = pix.tobytes()  # 获取PNG格式数据
            return Image.open(io.BytesIO(img_bytes))
    except:
        traceback.print_exc()
        return None
    # finally:
    #     os.remove(tmp_png)


#
def convert_word_to_image(word_path):
    tmp_pdf = os.path.join(tempfile.gettempdir(), str(uuid.uuid4()) + "_" + str(int(time.time())) + ".pdf")
    convert(word_path, tmp_pdf)
    image_date = convert_pdf_to_image(tmp_pdf)
    os.remove(tmp_pdf)
    return image_date


def convert_excel_to_image(excel_path):
    tmp_pdf = os.path.join(tempfile.gettempdir(), str(uuid.uuid4()) + "_" + str(int(time.time())) + ".pdf")
    wb = load_workbook(filename=excel_path, read_only=True)
    ws = wb.active
    pdf_buffer = BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=letter)
    for row in ws.iter_rows():
        for cell in row:
            c.drawString(cell.column * inch, (ws.max_row - cell.row + 1) * inch, str(cell.value))
    c.save()
    open(tmp_pdf, "wb").write(pdf_buffer.getvalue())
    image_date = convert_pdf_to_image(tmp_pdf)
    os.remove(tmp_pdf)
    return image_date


#

def is_txt_file(file_path_lower):
    return file_path_lower.endswith(".txt")


def is_pdf_file(file_path_lower):
    return file_path_lower.endswith(".pdf")


def is_doc_file(file_path_lower):
    return file_path_lower.endswith(".doc") or file_path_lower.endswith(".docx")


def is_image_file(file_path_lower):
    return file_path_lower[-3:] in ["png", "jpg", "bmp", "gif"] or file_path_lower[-4:] == "jpeg"


def is_excel_file(file_path_lower):
    return file_path_lower.endswith("xls") or file_path_lower.endswith("xlsx")


def file_preview(file_path: str):
    file_path_lower = file_path.lower()
    if is_txt_file(file_path_lower):
        return open(file_path, "r").readlines(20)
    if is_pdf_file(file_path_lower):
        return convert_pdf_to_image(file_path)
    if is_doc_file(file_path_lower):
        return convert_word_to_image(file_path)
    if is_excel_file(file_path_lower):
        return convert_excel_to_image(file_path)
    if is_image_file(file_path_lower):
        return Image.open(file_path)
    return None
